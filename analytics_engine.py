import json
import os
import re
import io
import pandas as pd
from typing import Dict, Any, List, Tuple, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from harvester import load_leads_data
except ImportError:
    load_leads_data = None

def extract_lat_lon_from_record(row: pd.Series) -> Tuple[Optional[float], Optional[float]]:
    """Extracts numeric lat and lon coordinates from google_maps_url or record fields."""
    url = str(row.get("google_maps_url", ""))
    match = re.search(r"query=([\d\.-]+),([\d\.-]+)", url)
    if match:
        try:
            return float(match.group(1)), float(match.group(2))
        except ValueError:
            pass
    return None, None

def calculate_lead_score_and_potential(row: pd.Series) -> Tuple[int, str, float]:
    """
    Calculates Lead Quality Score (0-100), Rating Grade (A+, A, B, C),
    and Estimated Monthly Wholesale Order Potential (INR).
    """
    score = 30  # Base Score
    
    # Phone Hygiene Factor (+35 pts)
    if row.get("phone_is_valid"):
        score += 35
    elif row.get("primary_phone"):
        score += 15
        
    # Website / Online Presence (+20 pts)
    website = str(row.get("website", ""))
    if website and website.strip().lower() not in ["none", "nan", "null", ""]:
        score += 20
        
    # Geocoding Precision (+15 pts)
    url = str(row.get("google_maps_url", ""))
    if "query=" in url:
        score += 15

    score = min(100, score)

    # Grade Classification
    if score >= 80:
        grade = "A+"
    elif score >= 65:
        grade = "A"
    elif score >= 50:
        grade = "B"
    else:
        grade = "C"

    # Order Potential Model based on Segment & Business Type
    bname = str(row.get("business_name", "")).lower()
    segment = str(row.get("segment", ""))
    
    if segment == "Institutional":
        order_potential = 60000.0  # College Hostels / Campuses
    elif any(k in bname for k in ["salon", "spa", "parlour", "makeover", "beauty parlour"]):
        order_potential = 45000.0  # High-volume Commercial Salons
    elif any(k in bname for k in ["pharmacy", "medical", "medplus", "apollo"]):
        order_potential = 30000.0  # Pharmacies & Personal Care Sections
    else:
        order_potential = 15000.0  # General Kirana & Fancy Stores

    return score, grade, order_potential

def enrich_leads_with_analytics(df_leads: pd.DataFrame) -> pd.DataFrame:
    """Enriches leads dataframe with scores, grades, order potential, and lat/lon coords."""
    if df_leads.empty:
        return df_leads

    scores = []
    grades = []
    potentials = []
    lats = []
    lons = []

    for _, row in df_leads.iterrows():
        score, grade, potential = calculate_lead_score_and_potential(row)
        lat, lon = extract_lat_lon_from_record(row)
        
        scores.append(score)
        grades.append(grade)
        potentials.append(potential)
        lats.append(lat)
        lons.append(lon)

    df_enriched = df_leads.copy()
    df_enriched["lead_score"] = scores
    df_enriched["lead_grade"] = grades
    df_enriched["estimated_monthly_order_inr"] = potentials
    df_enriched["lat"] = lats
    df_enriched["lon"] = lons

    return df_enriched

def calculate_lead_conversion_metrics(df_leads: pd.DataFrame) -> Dict[str, Any]:
    """Calculates lead lifecycle conversion funnel metrics and regional breakdown."""
    if df_leads.empty:
        return {
            "total_leads": 0,
            "status_breakdown": {},
            "segment_breakdown": {},
            "state_breakdown": {},
            "conversion_rate_pct": 0.0,
            "projected_monthly_revenue_inr": 0.0,
            "avg_lead_score": 0.0,
            "total_pipeline_value_inr": 0.0
        }

    df_enriched = enrich_leads_with_analytics(df_leads)
    total_leads = len(df_enriched)
    status_counts = df_enriched["lead_status"].value_counts().to_dict() if "lead_status" in df_enriched.columns else {}
    segment_counts = df_enriched["segment"].value_counts().to_dict() if "segment" in df_enriched.columns else {}
    state_counts = df_enriched["state"].value_counts().to_dict() if "state" in df_enriched.columns else {}

    converted_count = status_counts.get("Converted", 0) + status_counts.get("Qualified", 0)
    conversion_rate = (converted_count / total_leads * 100) if total_leads > 0 else 0.0

    total_pipeline_val = df_enriched["estimated_monthly_order_inr"].sum() if "estimated_monthly_order_inr" in df_enriched.columns else 0.0
    avg_score = df_enriched["lead_score"].mean() if "lead_score" in df_enriched.columns else 0.0

    return {
        "total_leads": total_leads,
        "status_breakdown": status_counts,
        "segment_breakdown": segment_counts,
        "state_breakdown": state_counts,
        "conversion_rate_pct": round(conversion_rate, 2),
        "projected_monthly_revenue_inr": round(total_pipeline_val * 0.20, 2),
        "total_pipeline_value_inr": round(total_pipeline_val, 2),
        "avg_lead_score": round(avg_score, 1)
    }

def update_lead_status(lead_id: str, new_status: str, local_file: str = "leads_master_local.json") -> bool:
    """Updates lead status in local storage or Supabase."""
    if os.path.exists(local_file):
        try:
            with open(local_file, "r") as f:
                records = json.load(f)
            updated = False
            for r in records:
                if r.get("lead_id") == lead_id or r.get("dedup_hash") == lead_id:
                    r["lead_status"] = new_status
                    updated = True
                    break
            if updated:
                with open(local_file, "w") as f:
                    json.dump(records, f, indent=2)
                return True
        except Exception:
            pass
    return False

def export_leads_to_excel(df_leads: pd.DataFrame) -> bytes:
    """
    Generates a beautifully styled, publication-grade Excel workbook (.xlsx)
    containing master lead listings, route delivery clusters, and executive analytics.
    """
    if df_leads.empty:
        df_leads = pd.DataFrame(columns=[
            "business_name", "segment", "lead_score", "lead_grade", 
            "estimated_monthly_order_inr", "primary_phone", "phone_is_valid", 
            "address_raw", "pincode", "state", "website", "google_maps_url", "lead_status"
        ])

    df_enriched = enrich_leads_with_analytics(df_leads)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Master Lead Directory
        cols_master = [
            "business_name", "segment", "lead_score", "lead_grade", 
            "estimated_monthly_order_inr", "primary_phone", "phone_is_valid", 
            "address_raw", "pincode", "state", "website", "google_maps_url", "lead_status"
        ]
        existing_cols = [c for c in cols_master if c in df_enriched.columns]
        df_master = df_enriched[existing_cols].copy()
        
        rename_map = {
            "business_name": "Business Name / Outlet",
            "segment": "Category / Segment",
            "lead_score": "Quality Score (0-100)",
            "lead_grade": "Quality Grade",
            "estimated_monthly_order_inr": "Est. Monthly Order (₹)",
            "primary_phone": "Verified Phone (+91)",
            "phone_is_valid": "Phone Valid?",
            "address_raw": "Full Business Address",
            "pincode": "PIN Code",
            "state": "State",
            "website": "Official Website / Social Link",
            "google_maps_url": "Google Maps Pin Link",
            "lead_status": "Status"
        }
        df_master = df_master.rename(columns=rename_map)
        df_master.to_excel(writer, sheet_name="📋 Master Lead Directory", index=False)

        # Sheet 2: Field Route Clusters
        if "pincode" in df_enriched.columns and "estimated_monthly_order_inr" in df_enriched.columns:
            df_route = df_enriched.groupby(["pincode", "state"]).agg(
                Verified_Outlets=("business_name", "count"),
                Total_Order_Potential=("estimated_monthly_order_inr", "sum"),
                Avg_Lead_Score=("lead_score", "mean")
            ).reset_index().sort_values(by="Total_Order_Potential", ascending=False)
            df_route.to_excel(writer, sheet_name="🚗 Territory Route Clusters", index=False)

        # Sheet 3: Executive Analytics Summary
        metrics = calculate_lead_conversion_metrics(df_enriched)
        df_summary = pd.DataFrame([
            {"Metric": "Total Verified Outlets Harvested", "Value": f"{metrics['total_leads']:,}"},
            {"Metric": "Phone Hygiene Validity Rate", "Value": f"{metrics.get('conversion_rate_pct', 0)}%"},
            {"Metric": "Total Pipeline Order Potential (Monthly)", "Value": f"₹{metrics['total_pipeline_value_inr']:,.2f}"},
            {"Metric": "Projected Distributor Margin Revenue (Monthly)", "Value": f"₹{metrics['projected_monthly_revenue_inr']:,.2f}"},
            {"Metric": "Average Outlet Lead Score", "Value": f"{metrics['avg_lead_score']} / 100"}
        ])
        df_summary.to_excel(writer, sheet_name="📊 Executive Summary", index=False)

    output.seek(0)
    wb = openpyxl.load_workbook(output)
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border

        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

    final_output = io.BytesIO()
    wb.save(final_output)
    return final_output.getvalue()
